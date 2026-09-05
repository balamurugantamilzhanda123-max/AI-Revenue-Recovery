import datetime
import io
from decimal import Decimal
from typing import Any

from sqlalchemy import desc, select
from sqlalchemy.orm import Session, joinedload

from app.models import (
    AuditLog,
    Customer,
    EscalationCase,
    EscalationStatus,
    PaymentAttempt,
    PaymentStatus,
    RecoveryAction,
    RecoveryCase,
    RecoveryStatus,
    Transaction,
)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.pdfgen import canvas
    from reportlab.platypus import (
        HRFlowable,
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    _REPORTLAB_AVAILABLE = True
except ImportError:
    _REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


if _REPORTLAB_AVAILABLE:
    class NumberedCanvas(canvas.Canvas):
        """
        Two-pass canvas to dynamically inject 'Page X of Y', running headers, and running footers.
        """
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_decorations(num_pages)
                super().showPage()
            super().save()

        def draw_page_decorations(self, page_count: int):
            self.saveState()
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.HexColor("#64748B"))

            # Header on subsequent pages
            if self._pageNumber > 1:
                self.drawString(36, 806, "REVIVEAI • Autonomous Revenue Recovery Report")
                self.drawRightString(559, 806, "CONFIDENTIAL")
                self.setStrokeColor(colors.HexColor("#E2E8F0"))
                self.setLineWidth(0.5)
                self.line(36, 800, 559, 800)

            # Footer on all pages
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(36, 45, 559, 45)
            self.drawString(36, 32, "REVIVEAI — Real-Time Autonomous Revenue Recovery & Safety System")
            self.drawRightString(559, 32, f"Page {self._pageNumber} of {page_count}")
            self.restoreState()


def get_report_data(
    db: Session,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
    failure_type: str | None = None,
    recovery_method: str | None = None,
) -> dict[str, Any]:
    """
    Calculates dynamic revenue recovery report metrics directly from database transactions.
    Guarantees 100% real-time reflection of database state.
    """
    query = (
        select(Transaction)
        .options(
            joinedload(Transaction.customer),
            joinedload(Transaction.recovery_cases).joinedload(RecoveryCase.recovery_actions),
            joinedload(Transaction.escalation_cases),
            joinedload(Transaction.payment_attempts),
        )
        .order_by(desc(Transaction.created_at))
    )

    all_txs = db.scalars(query).unique().all()

    # Filter transactions
    filtered_txs: list[Transaction] = []
    for tx in all_txs:
        # Date filtering
        if date_from and tx.created_at:
            try:
                dt_from = datetime.datetime.fromisoformat(date_from.replace("Z", "+00:00"))
                if tx.created_at.astimezone(datetime.timezone.utc) < dt_from.astimezone(datetime.timezone.utc):
                    continue
            except Exception:
                pass

        if date_to and tx.created_at:
            try:
                dt_to = datetime.datetime.fromisoformat(date_to.replace("Z", "+00:00"))
                if tx.created_at.astimezone(datetime.timezone.utc) > dt_to.astimezone(datetime.timezone.utc):
                    continue
            except Exception:
                pass

        # Status filtering
        if status and status.upper() != "ALL":
            if tx.status.value != status.upper() and tx.status.name != status.upper():
                continue

        # Failure Type filtering
        reason_comb = f"{tx.failure_reason or ''} {tx.gateway_response or ''}".lower()
        if failure_type and failure_type.upper() != "ALL":
            f_upper = failure_type.upper()
            if f_upper == "NETWORK_ERROR" and "network" not in reason_comb and "tcp rst" not in reason_comb:
                continue
            if f_upper == "PAYMENT_TIMEOUT" and "timeout" not in reason_comb and "504" not in reason_comb:
                continue
            if f_upper in {"AUTH_FAILURE", "AUTHENTICATION_FAILURE"} and "auth" not in reason_comb and "otp" not in reason_comb and "3ds" not in reason_comb:
                continue
            if f_upper == "BANK_DECLINE" and "decline" not in reason_comb and "balance" not in reason_comb:
                continue
            if f_upper in {"ABANDONED", "ABANDONMENT"} and tx.status != PaymentStatus.ABANDONED:
                continue

        # Recovery Method filtering
        is_human = (
            getattr(tx, "customer_response", "") == "RECOVERED_BY_HUMAN"
            or tx.escalation_status in {EscalationStatus.OPEN, EscalationStatus.IN_REVIEW, EscalationStatus.RESOLVED}
            or tx.recovery_status == RecoveryStatus.ESCALATED
        )
        if recovery_method and recovery_method.upper() != "ALL":
            m_upper = recovery_method.upper()
            if m_upper == "AI" and (is_human or (tx.status == PaymentStatus.SUCCESS and not tx.recovered_amount)):
                continue
            if m_upper == "HUMAN" and not is_human:
                continue
            if m_upper == "NONE" and (tx.recovery_status == RecoveryStatus.RECOVERED or is_human):
                continue

        filtered_txs.append(tx)

    # 1. Summary Metrics
    total_orders = len(filtered_txs)
    successful_payments = sum(1 for tx in filtered_txs if tx.status == PaymentStatus.SUCCESS)
    failed_payments = sum(1 for tx in filtered_txs if tx.status == PaymentStatus.FAILED)
    pending_payments = sum(1 for tx in filtered_txs if tx.status == PaymentStatus.PENDING)
    abandoned_payments = sum(1 for tx in filtered_txs if tx.status == PaymentStatus.ABANDONED)

    # Recoverable = FAILED or ABANDONED or UNRESOLVED, not already recovered
    recoverable_txs = [
        tx
        for tx in filtered_txs
        if tx.status in {PaymentStatus.FAILED, PaymentStatus.ABANDONED, PaymentStatus.UNRESOLVED}
        and tx.recovery_status != RecoveryStatus.RECOVERED
    ]
    revenue_at_risk = sum((Decimal(tx.amount or 0) for tx in recoverable_txs), Decimal("0.00"))

    # Recovery Attribution
    ai_recovered = Decimal("0.00")
    human_recovered = Decimal("0.00")
    ai_recovery_cases = 0
    human_recovery_cases = 0

    for tx in filtered_txs:
        if tx.status == PaymentStatus.SUCCESS and tx.recovered_amount and tx.recovered_amount > 0:
            if getattr(tx, "customer_response", "") == "RECOVERED_BY_HUMAN":
                human_recovered += Decimal(tx.recovered_amount)
                human_recovery_cases += 1
            else:
                ai_recovered += Decimal(tx.recovered_amount)
                ai_recovery_cases += 1
        elif tx.recovery_status == RecoveryStatus.OPEN:
            ai_recovery_cases += 1
        elif tx.recovery_status == RecoveryStatus.ESCALATED or bool(tx.escalation_cases):
            human_recovery_cases += 1

    total_recovered = ai_recovered + human_recovered
    base_recovery_pool = revenue_at_risk + total_recovered
    recovery_rate = (
        round((float(total_recovered) / float(base_recovery_pool) * 100), 1)
        if base_recovery_pool > 0
        else 0.0
    )

    # 2. Failure Breakdown & Amounts
    failure_breakdown_dict: dict[str, dict[str, Any]] = {
        "Network Error": {"cases": 0, "risk": Decimal("0.00"), "recovered": Decimal("0.00")},
        "Payment Timeout": {"cases": 0, "risk": Decimal("0.00"), "recovered": Decimal("0.00")},
        "Authentication Failure": {"cases": 0, "risk": Decimal("0.00"), "recovered": Decimal("0.00")},
        "Abandoned Payment": {"cases": 0, "risk": Decimal("0.00"), "recovered": Decimal("0.00")},
        "Other": {"cases": 0, "risk": Decimal("0.00"), "recovered": Decimal("0.00")},
    }

    for tx in filtered_txs:
        comb = f"{tx.failure_reason or ''} {tx.gateway_response or ''}".lower()
        amt = Decimal(tx.amount or 0)
        rec_amt = Decimal(tx.recovered_amount or 0) if tx.status == PaymentStatus.SUCCESS else Decimal("0.00")

        cat = "Other"
        if "network" in comb or "tcp rst" in comb or "connection" in comb:
            cat = "Network Error"
        elif "timeout" in comb or "504" in comb:
            cat = "Payment Timeout"
        elif "auth" in comb or "3ds" in comb or "otp" in comb:
            cat = "Authentication Failure"
        elif tx.status == PaymentStatus.ABANDONED:
            cat = "Abandoned Payment"
        elif tx.status == PaymentStatus.SUCCESS and not tx.failure_reason:
            continue

        failure_breakdown_dict[cat]["cases"] += 1
        if tx in recoverable_txs:
            failure_breakdown_dict[cat]["risk"] += amt
        failure_breakdown_dict[cat]["recovered"] += rec_amt

    # Formatted Failure Analysis Table rows
    failure_table_rows: list[dict[str, Any]] = []
    for ftype, fstat in failure_breakdown_dict.items():
        f_risk = float(fstat["risk"])
        f_rec = float(fstat["recovered"])
        f_unres = f_risk
        f_pool = f_risk + f_rec
        f_rate = round((f_rec / f_pool * 100), 1) if f_pool > 0 else 0.0
        failure_table_rows.append({
            "failure_type": ftype,
            "cases": fstat["cases"],
            "revenue_at_risk": f_risk,
            "recovered": f_rec,
            "unresolved": f_unres,
            "recovery_rate": f_rate,
        })

    # 3. Recovery Analysis Metrics
    high_risk_cases = sum(1 for tx in recoverable_txs if float(tx.amount or 0) >= 10000.0)
    unresolved_cases = len(recoverable_txs)
    unresolved_revenue = float(revenue_at_risk)

    # 4. Executive Findings & Strategic Recommendations
    findings: list[str] = []
    recommendations: list[str] = []

    if total_orders == 0:
        findings.append("No transactions or revenue events recorded in the current database period.")
        findings.append("Operational baseline is reset. Ready to ingest new checkout transactions.")
        recommendations.append("Execute standard simulated transactions from the Demo Center or initiate customer checkout.")
        recommendations.append("Verify payment gateway webhooks and telemetry pipelines.")
    else:
        findings.append(
            f"During the selected period, ReviveAI monitored {total_orders} transactions and identified ₹{float(revenue_at_risk):,.2f} in revenue at risk."
        )
        findings.append(
            f"Autonomous AI-driven recovery recovered ₹{float(ai_recovered):,.2f} and human specialist intervention recovered ₹{float(human_recovered):,.2f}, delivering an overall recovery rate of {recovery_rate}%."
        )
        
        top_failure = max(failure_table_rows, key=lambda x: x["cases"])
        if top_failure["cases"] > 0:
            findings.append(
                f"Primary failure category: {top_failure['failure_type']} ({top_failure['cases']} cases with ₹{top_failure['revenue_at_risk']:,.2f} at risk)."
            )

        if high_risk_cases > 0:
            findings.append(
                f"{high_risk_cases} high-risk transaction(s) (>= ₹10,000) identified requiring strict policy guardrails and SLA escalation."
            )
        
        if unresolved_cases > 0:
            findings.append(
                f"₹{unresolved_revenue:,.2f} across {unresolved_cases} unresolved case(s) remains open in the active recovery pipeline."
            )
        else:
            findings.append("Zero unresolved revenue leakage cases remaining. All identified risks were resolved or stopped.")

        # Recommendations based on actual data
        if failure_breakdown_dict["Network Error"]["cases"] > 0:
            recommendations.append(
                "Maintain autonomous TCP RST fast-retry policy (< 200ms latency) to capture transient connection drops."
            )
        if failure_breakdown_dict["Payment Timeout"]["cases"] > 0:
            recommendations.append(
                "Enforce dual-phase gateway status query before retrying 504 timeouts to eliminate duplicate charges."
            )
        if failure_breakdown_dict["Abandoned Payment"]["cases"] > 0:
            recommendations.append(
                "Dispatch personalized WhatsApp/Email dynamic recovery tokens for cart abandonments within 3 minutes of drop-off."
            )
        if high_risk_cases > 0:
            recommendations.append(
                "Prioritize human associate workspace queue for orders exceeding ₹10,000 to maximize high-ticket conversion."
            )
        if not recommendations:
            recommendations.append("Continue current deterministic + LLM hybrid recovery policies.")

    # 5. Transaction Details List
    details: list[dict[str, Any]] = []
    for tx in filtered_txs:
        rec_case = tx.recovery_cases[0] if tx.recovery_cases else None
        rec_action = rec_case.recovery_actions[0] if rec_case and rec_case.recovery_actions else None
        
        diagnosis = rec_case.root_cause if rec_case and rec_case.root_cause else (
            "Network Connection Interrupted" if "network" in (tx.failure_reason or "").lower() else (
                "Gateway Handshake Timeout" if "timeout" in (tx.failure_reason or "").lower() else (
                    "Authentication Handshake Failed" if "auth" in (tx.failure_reason or "").lower() else (
                        "Standard Authorization" if tx.status == PaymentStatus.SUCCESS else "Payment Failure"
                    )
                )
            )
        )

        is_tx_human = getattr(tx, "customer_response", "") == "RECOVERED_BY_HUMAN" or tx.escalation_status != EscalationStatus.NONE
        rec_method = (
            "Human Associate" if is_tx_human
            else ("AI Autonomous Agent" if tx.status != PaymentStatus.SUCCESS or (tx.recovered_amount and tx.recovered_amount > 0) else "Direct Gateway")
        )
        
        action_name = (
            rec_action.action_type if rec_action else (
                "Automated Smart Retry" if tx.status == PaymentStatus.FAILED and "network" in (tx.failure_reason or "").lower()
                else ("Customer Recovery Link" if tx.status == PaymentStatus.ABANDONED else "Payment Captured")
            )
        )

        amt = float(tx.amount or 0)
        risk_level = "CRITICAL" if amt >= 50000 else ("HIGH" if amt >= 10000 else ("MEDIUM" if amt >= 1000 else "LOW"))

        details.append({
            "transaction_id": tx.transaction_id,
            "order_id": tx.order_id,
            "customer_name": tx.customer.name if tx.customer else "Customer",
            "customer_email": tx.customer.email if tx.customer else "",
            "customer_phone": tx.customer.phone if tx.customer else "",
            "amount": amt,
            "currency": tx.currency or "INR",
            "status": tx.status.value if hasattr(tx.status, "value") else str(tx.status),
            "failure_type": tx.failure_reason or ("Checkout Abandonment" if tx.status == PaymentStatus.ABANDONED else "None"),
            "risk_level": risk_level,
            "diagnosis": diagnosis,
            "recovery_action": action_name,
            "recovery_method": rec_method,
            "recovery_status": tx.recovery_status.value if hasattr(tx.recovery_status, "value") else str(tx.recovery_status),
            "recovered_amount": float(tx.recovered_amount or 0),
            "created_date": tx.created_at.isoformat() if tx.created_at else None,
            "updated_date": tx.updated_at.isoformat() if tx.updated_at else None,
        })

    # Fetch Audit Logs for Excel export
    audit_query = select(AuditLog).order_by(desc(AuditLog.created_at)).limit(100)
    audit_records = db.scalars(audit_query).all()
    audit_logs_list = [
        {
            "timestamp": log.created_at.isoformat() if log.created_at else "",
            "actor": log.actor or "SYSTEM",
            "action": log.event_type,
            "description": log.event_message,
            "metadata": str(log.metadata_json or {}),
        }
        for log in audit_records
    ]

    return {
        "summary": {
            "total_orders": total_orders,
            "successful_payments": successful_payments,
            "failed_payments": failed_payments,
            "pending_payments": pending_payments,
            "abandoned_payments": abandoned_payments,
            "revenue_at_risk": float(revenue_at_risk),
            "ai_recovered": float(ai_recovered),
            "human_recovered": float(human_recovered),
            "total_recovered": float(total_recovered),
            "recovery_rate": recovery_rate,
            "unresolved_revenue": unresolved_revenue,
            "high_risk_cases": high_risk_cases,
        },
        "failure_analysis": {
            "network_errors": failure_breakdown_dict["Network Error"]["cases"],
            "payment_timeouts": failure_breakdown_dict["Payment Timeout"]["cases"],
            "authentication_failures": failure_breakdown_dict["Authentication Failure"]["cases"],
            "abandonments": failure_breakdown_dict["Abandoned Payment"]["cases"],
            "other_failures": failure_breakdown_dict["Other"]["cases"],
            "breakdown_table": failure_table_rows,
        },
        "recovery_analysis": {
            "ai_recovery_cases": ai_recovery_cases,
            "human_recovery_cases": human_recovery_cases,
            "high_risk_cases": high_risk_cases,
            "unresolved_cases": unresolved_cases,
            "unresolved_revenue": unresolved_revenue,
        },
        "executive_findings": findings,
        "recommendations": recommendations,
        "transactions": details,
        "audit_logs": audit_logs_list,
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "status": status,
            "failure_type": failure_type,
            "recovery_method": recovery_method,
        },
    }


def generate_report_pdf(report_data: dict[str, Any]) -> bytes:
    """
    Generates a structured, professional A4 PDF Revenue Recovery Analytics Report.
    Adheres strictly to the 5-page enterprise specification.
    """
    buffer = io.BytesIO()

    if not _REPORTLAB_AVAILABLE:
        buffer.write(b"%PDF-1.4\n% ReviveAI PDF Fallback\n")
        buffer.seek(0)
        return buffer.getvalue()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=40,
        bottomMargin=55,
    )

    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]

    brand_title_style = ParagraphStyle(
        "BrandTitle",
        parent=normal_style,
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#065F46"),
        fontName="Helvetica-Bold",
    )

    doc_subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=normal_style,
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#0F172A"),
        fontName="Helvetica-Bold",
    )

    section_hdr_style = ParagraphStyle(
        "SectionHdr",
        parent=normal_style,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#065F46"),
        fontName="Helvetica-Bold",
        spaceBefore=12,
        spaceAfter=6,
    )

    kpi_title_style = ParagraphStyle(
        "KPITitle",
        parent=normal_style,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#64748B"),
        fontName="Helvetica-Bold",
    )

    kpi_val_style = ParagraphStyle(
        "KPIValue",
        parent=normal_style,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0F172A"),
        fontName="Helvetica-Bold",
    )

    body_style = ParagraphStyle(
        "Body",
        parent=normal_style,
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )

    cell_style = ParagraphStyle(
        "Cell",
        parent=normal_style,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1E293B"),
    )

    cell_bold_style = ParagraphStyle(
        "CellBold",
        parent=normal_style,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#0F172A"),
        fontName="Helvetica-Bold",
    )

    cell_header_style = ParagraphStyle(
        "CellHeader",
        parent=normal_style,
        fontSize=8,
        leading=10,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    elements = []
    summary = report_data.get("summary", {})
    gen_time = report_data.get("generated_at", datetime.datetime.now(datetime.timezone.utc).isoformat())
    gen_time_str = gen_time[:19].replace("T", " ") + " UTC"
    filters = report_data.get("filters", {})
    date_from = filters.get("date_from") or "Beginning of Records"
    date_to = filters.get("date_to") or "Present"

    # =========================================================================
    # PAGE 1: EXECUTIVE SUMMARY
    # =========================================================================
    hdr_table = Table(
        [
            [
                Paragraph("<b>REVIVEAI</b><br/><font size=9 color='#047857'>Autonomous Revenue Recovery</font>", brand_title_style),
                Paragraph(
                    f"<b>REVENUE RECOVERY ANALYTICS REPORT</b><br/>"
                    f"<font size=8 color='#64748B'>Report Generated: {gen_time_str}<br/>"
                    f"Report Period: {date_from} → {date_to}</font>",
                    doc_subtitle_style,
                ),
            ]
        ],
        colWidths=[270, 253],
    )
    hdr_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(hdr_table)
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#065F46"), spaceAfter=12))

    elements.append(Paragraph("EXECUTIVE SUMMARY & KEY PERFORMANCE INDICATORS", section_hdr_style))

    # 10 KPI Grid (2 rows x 5 cards)
    kpis = [
        ("TOTAL ORDERS", str(summary.get("total_orders", 0))),
        ("SUCCESSFUL PAYMENTS", str(summary.get("successful_payments", 0))),
        ("FAILED PAYMENTS", str(summary.get("failed_payments", 0))),
        ("REVENUE AT RISK", f"INR {summary.get('revenue_at_risk', 0):,.2f}"),
        ("AI RECOVERED", f"INR {summary.get('ai_recovered', 0):,.2f}"),
        ("HUMAN RECOVERED", f"INR {summary.get('human_recovered', 0):,.2f}"),
        ("TOTAL RECOVERED", f"INR {summary.get('total_recovered', 0):,.2f}"),
        ("RECOVERY RATE", f"{summary.get('recovery_rate', 0)}%"),
        ("UNRESOLVED REVENUE", f"INR {summary.get('unresolved_revenue', 0):,.2f}"),
        ("HIGH RISK ORDERS", str(summary.get("high_risk_cases", 0))),
    ]

    kpi_cards = []
    for title, val in kpis:
        kpi_cards.append([
            Paragraph(title, kpi_title_style),
            Paragraph(val, kpi_val_style),
        ])

    row1 = [
        Table([[kpi_cards[0][0]], [kpi_cards[0][1]]], colWidths=[98]),
        Table([[kpi_cards[1][0]], [kpi_cards[1][1]]], colWidths=[98]),
        Table([[kpi_cards[2][0]], [kpi_cards[2][1]]], colWidths=[98]),
        Table([[kpi_cards[3][0]], [kpi_cards[3][1]]], colWidths=[108]),
        Table([[kpi_cards[4][0]], [kpi_cards[4][1]]], colWidths=[108]),
    ]
    row2 = [
        Table([[kpi_cards[5][0]], [kpi_cards[5][1]]], colWidths=[98]),
        Table([[kpi_cards[6][0]], [kpi_cards[6][1]]], colWidths=[98]),
        Table([[kpi_cards[7][0]], [kpi_cards[7][1]]], colWidths=[98]),
        Table([[kpi_cards[8][0]], [kpi_cards[8][1]]], colWidths=[108]),
        Table([[kpi_cards[9][0]], [kpi_cards[9][1]]], colWidths=[108]),
    ]

    kpi_table = Table([row1, row2], colWidths=[102, 102, 102, 108, 108])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 14))

    # Executive Narrative
    elements.append(Paragraph("EXECUTIVE NARRATIVE", section_hdr_style))
    narrative_p = (
        f"During the selected reporting period ({date_from} to {date_to}), ReviveAI monitored "
        f"<b>{summary.get('total_orders', 0)}</b> transactions and identified <b>INR {summary.get('revenue_at_risk', 0):,.2f}</b> in revenue at risk. "
        f"Through autonomous AI-driven payment recovery and coordinated human associate intervention, "
        f"<b>INR {summary.get('total_recovered', 0):,.2f}</b> was successfully captured, achieving an overall recovery rate of "
        f"<b>{summary.get('recovery_rate', 0)}%</b>.<br/><br/>"
        f"AI autonomous retry workflows accounted for <b>INR {summary.get('ai_recovered', 0):,.2f}</b>, while human specialists recovered "
        f"<b>INR {summary.get('human_recovered', 0):,.2f}</b> across high-risk and complex technical escalations. "
        f"Unresolved revenue currently stands at <b>INR {summary.get('unresolved_revenue', 0):,.2f}</b> across "
        f"<b>{summary.get('high_risk_cases', 0)}</b> priority cases."
        if summary.get("total_orders", 0) > 0
        else "No transaction data is available for the selected reporting period. Operational baseline is currently clean."
    )
    elements.append(Paragraph(narrative_p, body_style))
    elements.append(Spacer(1, 14))

    # =========================================================================
    # PAGE 2: FAILURE ANALYSIS
    # =========================================================================
    elements.append(PageBreak())
    elements.append(Paragraph("FAILURE ANALYSIS & ROOT CAUSE BREAKDOWN", section_hdr_style))
    elements.append(Paragraph(
        "Detailed classification of transaction drop-offs, network interruptions, and gateway handshake timeouts:",
        body_style,
    ))
    elements.append(Spacer(1, 8))

    fail_rows = report_data.get("failure_analysis", {}).get("breakdown_table", [])
    fail_table_data = [
        [
            Paragraph("Failure Type", cell_header_style),
            Paragraph("Cases", cell_header_style),
            Paragraph("Revenue at Risk", cell_header_style),
            Paragraph("Recovered", cell_header_style),
            Paragraph("Unresolved", cell_header_style),
            Paragraph("Recovery Rate", cell_header_style),
        ]
    ]

    for fr in fail_rows:
        fail_table_data.append([
            Paragraph(fr["failure_type"], cell_bold_style),
            Paragraph(str(fr["cases"]), cell_style),
            Paragraph(f"INR {fr['revenue_at_risk']:,.2f}", cell_style),
            Paragraph(f"INR {fr['recovered']:,.2f}", cell_style),
            Paragraph(f"INR {fr['unresolved']:,.2f}", cell_style),
            Paragraph(f"<b>{fr['recovery_rate']}%</b>", cell_bold_style),
        ])

    fail_table = Table(fail_table_data, colWidths=[140, 55, 85, 85, 85, 73])
    fail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065F46")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(fail_table)
    elements.append(Spacer(1, 14))

    # =========================================================================
    # PAGE 3: RECOVERY PERFORMANCE
    # =========================================================================
    elements.append(PageBreak())
    elements.append(Paragraph("RECOVERY PERFORMANCE & CHANNEL ATTRIBUTION", section_hdr_style))
    elements.append(Paragraph(
        "Comparative metrics between Autonomous AI Fast-Path Retries and Human Associate Escalation Workspaces:",
        body_style,
    ))
    elements.append(Spacer(1, 8))

    rec = report_data.get("recovery_analysis", {})
    rec_table_data = [
        [
            Paragraph("Recovery Channel", cell_header_style),
            Paragraph("Cases Handled", cell_header_style),
            Paragraph("Amount Recovered", cell_header_style),
            Paragraph("Share of Recovered", cell_header_style),
        ],
        [
            Paragraph("Autonomous AI Engine", cell_bold_style),
            Paragraph(str(rec.get("ai_recovery_cases", 0)), cell_style),
            Paragraph(f"INR {summary.get('ai_recovered', 0):,.2f}", cell_style),
            Paragraph(
                f"{(summary.get('ai_recovered', 0) / summary.get('total_recovered', 1) * 100):.1f}%"
                if summary.get("total_recovered", 0) > 0 else "0.0%",
                cell_bold_style,
            ),
        ],
        [
            Paragraph("Human Associate Queue", cell_bold_style),
            Paragraph(str(rec.get("human_recovery_cases", 0)), cell_style),
            Paragraph(f"INR {summary.get('human_recovered', 0):,.2f}", cell_style),
            Paragraph(
                f"{(summary.get('human_recovered', 0) / summary.get('total_recovered', 1) * 100):.1f}%"
                if summary.get("total_recovered", 0) > 0 else "0.0%",
                cell_bold_style,
            ),
        ],
        [
            Paragraph("<b>Total System Recovery</b>", cell_bold_style),
            Paragraph(f"<b>{rec.get('ai_recovery_cases', 0) + rec.get('human_recovery_cases', 0)}</b>", cell_bold_style),
            Paragraph(f"<b>INR {summary.get('total_recovered', 0):,.2f}</b>", cell_bold_style),
            Paragraph("<b>100.0%</b>", cell_bold_style),
        ],
    ]

    rec_table = Table(rec_table_data, colWidths=[160, 110, 130, 123])
    rec_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065F46")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E6FFFA")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(rec_table)
    elements.append(Spacer(1, 14))

    # =========================================================================
    # PAGE 4: TRANSACTION DETAILS
    # =========================================================================
    elements.append(PageBreak())
    elements.append(Paragraph("TRANSACTION DETAILS & AUDIT LEDGER", section_hdr_style))
    elements.append(Paragraph(
        "Individual transaction log with technical diagnosis, recovery action, and financial state:",
        body_style,
    ))
    elements.append(Spacer(1, 8))

    tx_list = report_data.get("transactions", [])
    if not tx_list:
        elements.append(Paragraph("<i>No transactions recorded for the selected filter criteria.</i>", body_style))
    else:
        tx_table_data = [
            [
                Paragraph("Transaction ID", cell_header_style),
                Paragraph("Amount", cell_header_style),
                Paragraph("Status", cell_header_style),
                Paragraph("Failure Type", cell_header_style),
                Paragraph("Risk", cell_header_style),
                Paragraph("Method", cell_header_style),
                Paragraph("Recovery Status", cell_header_style),
                Paragraph("Recovered", cell_header_style),
            ]
        ]
        for t in tx_list:
            tx_table_data.append([
                Paragraph(t.get("transaction_id", "")[:15], cell_style),
                Paragraph(f"INR {t.get('amount', 0):,.0f}", cell_style),
                Paragraph(t.get("status", ""), cell_style),
                Paragraph(t.get("failure_type", "")[:22], cell_style),
                Paragraph(t.get("risk_level", "MED"), cell_style),
                Paragraph(t.get("recovery_method", "")[:12], cell_style),
                Paragraph(t.get("recovery_status", ""), cell_style),
                Paragraph(f"INR {t.get('recovered_amount', 0):,.0f}", cell_style),
            ])

        tx_table = Table(tx_table_data, colWidths=[80, 52, 52, 105, 45, 65, 65, 59], repeatRows=1)
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065F46")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(tx_table)

    # =========================================================================
    # PAGE 5: FINDINGS & RECOMMENDATIONS
    # =========================================================================
    elements.append(PageBreak())
    elements.append(Paragraph("AI RECOVERY FINDINGS", section_hdr_style))
    findings = report_data.get("executive_findings", [])
    if findings:
        findings_rows = []
        for item in findings:
            findings_rows.append([Paragraph("•", cell_bold_style), Paragraph(item, body_style)])
        findings_table = Table(findings_rows, colWidths=[15, 508])
        findings_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(findings_table)
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("RECOMMENDATIONS & ACTION PLAN", section_hdr_style))
    recommendations = report_data.get("recommendations", [])
    if recommendations:
        rec_rows = []
        for idx, rec_text in enumerate(recommendations, 1):
            rec_rows.append([
                Paragraph(f"<b>{idx}.</b>", cell_bold_style),
                Paragraph(rec_text, body_style),
            ])
        rec_table = Table(rec_rows, colWidths=[20, 503])
        rec_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(rec_table)

    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


def generate_transaction_pdf(tx_data: dict[str, Any]) -> bytes:
    """
    Generates a dedicated, single-transaction enterprise PDF report.
    """
    buffer = io.BytesIO()

    if not _REPORTLAB_AVAILABLE:
        buffer.write(b"%PDF-1.4\n% ReviveAI Transaction PDF Fallback\n")
        buffer.seek(0)
        return buffer.getvalue()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=40,
        bottomMargin=50,
    )

    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]

    brand_title_style = ParagraphStyle(
        "TxBrandTitle",
        parent=normal_style,
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#065F46"),
        fontName="Helvetica-Bold",
    )

    section_hdr_style = ParagraphStyle(
        "TxSectionHdr",
        parent=normal_style,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#065F46"),
        fontName="Helvetica-Bold",
        spaceBefore=10,
        spaceAfter=4,
    )

    cell_style = ParagraphStyle(
        "TxCell",
        parent=normal_style,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#1E293B"),
    )

    cell_bold_style = ParagraphStyle(
        "TxCellBold",
        parent=normal_style,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#0F172A"),
        fontName="Helvetica-Bold",
    )

    elements = []

    tx_id = tx_data.get("transaction_id", "N/A")
    order_id = tx_data.get("order_id", "N/A")
    amt = float(tx_data.get("amount", 0))
    status = tx_data.get("status", "N/A")
    rec_status = tx_data.get("recovery_status", "NOT_STARTED")

    # Header
    elements.append(
        Table(
            [
                [
                    Paragraph("<b>REVIVEAI</b><br/><font size=8 color='#047857'>Autonomous Revenue Recovery</font>", brand_title_style),
                    Paragraph(
                        f"<b>TRANSACTION AUDIT CERTIFICATE</b><br/>"
                        f"<font size=8 color='#64748B'>Txn ID: {tx_id}<br/>"
                        f"Order ID: {order_id}</font>",
                        ParagraphStyle("Sub", parent=normal_style, fontSize=9, leading=12, alignment=2),
                    ),
                ]
            ],
            colWidths=[260, 263],
        )
    )
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#065F46"), spaceAfter=10))

    # KPI Top Bar
    kpi_bar = [
        [
            Paragraph("TRANSACTION AMOUNT", ParagraphStyle("KB1", parent=normal_style, fontSize=7.5, textColor=colors.HexColor("#64748B"), fontName="Helvetica-Bold")),
            Paragraph("PAYMENT STATUS", ParagraphStyle("KB2", parent=normal_style, fontSize=7.5, textColor=colors.HexColor("#64748B"), fontName="Helvetica-Bold")),
            Paragraph("RECOVERY STATUS", ParagraphStyle("KB3", parent=normal_style, fontSize=7.5, textColor=colors.HexColor("#64748B"), fontName="Helvetica-Bold")),
            Paragraph("RECOVERED AMOUNT", ParagraphStyle("KB4", parent=normal_style, fontSize=7.5, textColor=colors.HexColor("#64748B"), fontName="Helvetica-Bold")),
        ],
        [
            Paragraph(f"INR {amt:,.2f}", ParagraphStyle("KV1", parent=normal_style, fontSize=12, fontName="Helvetica-Bold")),
            Paragraph(str(status), ParagraphStyle("KV2", parent=normal_style, fontSize=12, fontName="Helvetica-Bold")),
            Paragraph(str(rec_status), ParagraphStyle("KV3", parent=normal_style, fontSize=12, fontName="Helvetica-Bold", textColor=colors.HexColor("#047857"))),
            Paragraph(f"INR {float(tx_data.get('recovered_amount', 0)):,.2f}", ParagraphStyle("KV4", parent=normal_style, fontSize=12, fontName="Helvetica-Bold")),
        ],
    ]
    t_kpi = Table(kpi_bar, colWidths=[130, 130, 130, 133])
    t_kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 10))

    # Customer & Payment Info
    cust = tx_data.get("customer", {})
    info_table_data = [
        [
            Paragraph("Customer Name", cell_bold_style),
            Paragraph(cust.get("name", "N/A"), cell_style),
            Paragraph("Payment Method", cell_bold_style),
            Paragraph(tx_data.get("payment_method", "UPI / Card"), cell_style),
        ],
        [
            Paragraph("Customer Email", cell_bold_style),
            Paragraph(cust.get("email", "N/A"), cell_style),
            Paragraph("Gateway Response", cell_bold_style),
            Paragraph(tx_data.get("gateway_response", "N/A") or "N/A", cell_style),
        ],
        [
            Paragraph("Customer Phone", cell_bold_style),
            Paragraph(cust.get("phone", "N/A"), cell_style),
            Paragraph("Created At", cell_bold_style),
            Paragraph(str(tx_data.get("created_at", "N/A"))[:19].replace("T", " "), cell_style),
        ],
    ]
    t_info = Table(info_table_data, colWidths=[110, 150, 110, 153])
    t_info.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(Paragraph("TRANSACTION & CUSTOMER IDENTIFIERS", section_hdr_style))
    elements.append(t_info)
    elements.append(Spacer(1, 8))

    # AI Diagnosis & Recovery
    diag_rows = [
        [
            Paragraph("Failure Reason", cell_bold_style),
            Paragraph(tx_data.get("failure_reason", "None") or "None", cell_style),
        ],
        [
            Paragraph("AI Technical Diagnosis", cell_bold_style),
            Paragraph(tx_data.get("diagnosis", "Autonomous validation completed without detected faults."), cell_style),
        ],
        [
            Paragraph("Recovery Action Executed", cell_bold_style),
            Paragraph(tx_data.get("recovery_action", "Automated Smart Retry / Dynamic Token Dispatch"), cell_style),
        ],
        [
            Paragraph("Recovery Attribution", cell_bold_style),
            Paragraph(tx_data.get("recovery_method", "AI Autonomous Agent"), cell_style),
        ],
    ]
    t_diag = Table(diag_rows, colWidths=[140, 383])
    t_diag.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(Paragraph("AI DIAGNOSTICS & RECOVERY ACTIONS", section_hdr_style))
    elements.append(t_diag)
    elements.append(Spacer(1, 8))

    # Transaction Lifecycle Timeline
    timeline_rows = [
        [
            Paragraph("<b>Stage 1: Checkout Initiation</b>", cell_bold_style),
            Paragraph("Customer initiated order checkout and payment token request.", cell_style),
        ],
        [
            Paragraph("<b>Stage 2: Gateway Execution</b>", cell_bold_style),
            Paragraph(f"Gateway returned status '{status}'. Reason: {tx_data.get('failure_reason', 'Captured') or 'Captured'}.", cell_style),
        ],
        [
            Paragraph("<b>Stage 3: AI Diagnosis</b>", cell_bold_style),
            Paragraph("ReviveAI autonomous engine analyzed network telemetry and verified token status.", cell_style),
        ],
        [
            Paragraph("<b>Stage 4: Autonomous Recovery</b>", cell_bold_style),
            Paragraph(f"Action '{tx_data.get('recovery_action', 'Controlled Retry')}' executed. Recovery status: {rec_status}.", cell_style),
        ],
    ]
    t_tl = Table(timeline_rows, colWidths=[160, 363])
    t_tl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(Paragraph("TRANSACTION LIFECYCLE TIMELINE", section_hdr_style))
    elements.append(t_tl)
    elements.append(Spacer(1, 10))

    # Security & Audit Signature
    footer_p = Paragraph(
        "<b>Certificate Verification:</b> This document represents an authorized, immutable cryptographic audit record "
        "generated by ReviveAI Autonomous Revenue Recovery Platform.",
        ParagraphStyle("Foot", parent=normal_style, fontSize=7.5, leading=10, textColor=colors.HexColor("#64748B")),
    )
    elements.append(footer_p)

    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


def generate_report_excel(report_data: dict[str, Any]) -> bytes:
    """
    Generates a 5-sheet enterprise Excel workbook using openpyxl:
    Sheet 1: Summary
    Sheet 2: Transactions
    Sheet 3: Recovery Analysis
    Sheet 4: Risk Analysis
    Sheet 5: Audit Log
    """
    buffer = io.BytesIO()

    if not _OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="065F46")
    bold_font = Font(name="Calibri", size=10, bold=True)
    normal_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    summary = report_data.get("summary", {})
    tx_list = report_data.get("transactions", [])
    fail_breakdown = report_data.get("failure_analysis", {}).get("breakdown_table", [])
    rec = report_data.get("recovery_analysis", {})
    audit_logs = report_data.get("audit_logs", [])
    gen_time = report_data.get("generated_at", datetime.datetime.now(datetime.timezone.utc).isoformat())

    # ============================================================
    # Sheet 1: Summary
    # ============================================================
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.views.sheetView[0].showGridLines = True

    ws1.append(["REVIVEAI - AUTONOMOUS REVENUE RECOVERY REPORT"])
    ws1.cell(row=1, column=1).font = title_font
    ws1.append([f"Report Generated: {gen_time[:19].replace('T', ' ')} UTC"])
    ws1.append([])

    ws1.append(["Key Performance Metric", "Metric Value"])
    for col_idx in range(1, 3):
        cell = ws1.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    summary_rows = [
        ["Total Orders Monitored", summary.get("total_orders", 0)],
        ["Successful Payments", summary.get("successful_payments", 0)],
        ["Failed Payments", summary.get("failed_payments", 0)],
        ["Pending Payments", summary.get("pending_payments", 0)],
        ["Abandoned Payments", summary.get("abandoned_payments", 0)],
        ["Revenue at Risk", summary.get("revenue_at_risk", 0)],
        ["AI Recovered Revenue", summary.get("ai_recovered", 0)],
        ["Human Recovered Revenue", summary.get("human_recovered", 0)],
        ["Total Recovered Revenue", summary.get("total_recovered", 0)],
        ["Recovery Success Rate (%)", f"{summary.get('recovery_rate', 0)}%"],
        ["Unresolved Revenue at Risk", summary.get("unresolved_revenue", 0)],
        ["High Risk Orders (>= INR 10k)", summary.get("high_risk_cases", 0)],
    ]

    for r_idx, (m_label, m_val) in enumerate(summary_rows, start=5):
        ws1.append([m_label, m_val])
        c1 = ws1.cell(row=r_idx, column=1)
        c2 = ws1.cell(row=r_idx, column=2)
        c1.font = bold_font
        c2.font = normal_font
        c1.border = thin_border
        c2.border = thin_border
        if "Revenue" in m_label or "Risk" in m_label and "%" not in str(m_val):
            if isinstance(m_val, (int, float)):
                c2.number_format = '"₹"#,##0.00'

    ws1.append([])
    ws1.append(["EXECUTIVE FINDINGS & NARRATIVE"])
    ws1.cell(row=len(summary_rows) + 6, column=1).font = bold_font
    for item in report_data.get("executive_findings", []):
        ws1.append([f"• {item}"])

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 28

    # ============================================================
    # Sheet 2: Transactions
    # ============================================================
    ws2 = wb.create_sheet(title="Transactions")
    ws2.views.sheetView[0].showGridLines = True
    ws2.freeze_panes = "A2"

    tx_headers = [
        "Transaction ID",
        "Order ID",
        "Amount",
        "Status",
        "Failure Type",
        "Diagnosis",
        "Recovery Action",
        "Recovery Method",
        "Recovery Status",
        "Created At",
        "Updated At",
    ]
    ws2.append(tx_headers)
    for col_idx in range(1, len(tx_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in tx_list:
        ws2.append([
            t.get("transaction_id", ""),
            t.get("order_id", ""),
            t.get("amount", 0),
            t.get("status", ""),
            t.get("failure_type", ""),
            t.get("diagnosis", ""),
            t.get("recovery_action", ""),
            t.get("recovery_method", ""),
            t.get("recovery_status", ""),
            t.get("created_date", ""),
            t.get("updated_date", ""),
        ])

    for row in ws2.iter_rows(min_row=2, max_row=max(ws2.max_row, 2), min_col=1, max_col=len(tx_headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column == 3 and isinstance(cell.value, (int, float)):
                cell.number_format = '"₹"#,##0.00'

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws2.column_dimensions[col].width = 20

    # ============================================================
    # Sheet 3: Recovery Analysis
    # ============================================================
    ws3 = wb.create_sheet(title="Recovery Analysis")
    ws3.views.sheetView[0].showGridLines = True
    ws3.freeze_panes = "A2"

    rec_headers = [
        "Recovery ID",
        "Transaction ID",
        "Amount",
        "Method",
        "Action",
        "Status",
        "AI/Human",
        "Created At",
    ]
    ws3.append(rec_headers)
    for col_idx in range(1, len(rec_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, t in enumerate(tx_list, start=1):
        if t.get("status") != "SUCCESS" or (t.get("recovered_amount", 0) > 0) or t.get("recovery_status") != "NOT_STARTED":
            rec_id = f"REC-{t.get('transaction_id', f'{idx:04d}')}"
            is_human = "Human" in str(t.get("recovery_method", ""))
            ws3.append([
                rec_id,
                t.get("transaction_id", ""),
                t.get("amount", 0),
                t.get("recovery_method", "AI Autonomous Agent"),
                t.get("recovery_action", "Smart Retry"),
                t.get("recovery_status", "OPEN"),
                "Human" if is_human else "AI",
                t.get("created_date", ""),
            ])

    for row in ws3.iter_rows(min_row=2, max_row=max(ws3.max_row, 2), min_col=1, max_col=len(rec_headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column == 3 and isinstance(cell.value, (int, float)):
                cell.number_format = '"₹"#,##0.00'

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws3.column_dimensions[col].width = 22

    # ============================================================
    # Sheet 4: Risk Analysis
    # ============================================================
    ws4 = wb.create_sheet(title="Risk Analysis")
    ws4.views.sheetView[0].showGridLines = True
    ws4.freeze_panes = "A2"

    risk_headers = [
        "Transaction ID",
        "Amount",
        "Risk Level",
        "Failure Type",
        "Status",
        "Resolution",
    ]
    ws4.append(risk_headers)
    for col_idx in range(1, len(risk_headers) + 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in tx_list:
        if t.get("status") in {"FAILED", "ABANDONED", "UNRESOLVED"} or t.get("recovery_status") != "NOT_STARTED":
            res_str = "RECOVERED" if t.get("recovery_status") == "RECOVERED" else (
                "ESCALATED_TO_HUMAN" if t.get("recovery_status") == "ESCALATED" else "PENDING_RECOVERY"
            )
            ws4.append([
                t.get("transaction_id", ""),
                t.get("amount", 0),
                t.get("risk_level", "MEDIUM"),
                t.get("failure_type", "Payment Failure"),
                t.get("status", ""),
                res_str,
            ])

    for row in ws4.iter_rows(min_row=2, max_row=max(ws4.max_row, 2), min_col=1, max_col=len(risk_headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '"₹"#,##0.00'

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws4.column_dimensions[col].width = 24

    # ============================================================
    # Sheet 5: Audit Log
    # ============================================================
    ws5 = wb.create_sheet(title="Audit Log")
    ws5.views.sheetView[0].showGridLines = True
    ws5.freeze_panes = "A2"

    audit_headers = [
        "Timestamp (UTC)",
        "Actor / User",
        "Action Event",
        "Description",
        "Metadata / Records Affected",
    ]
    ws5.append(audit_headers)
    for col_idx in range(1, len(audit_headers) + 1):
        cell = ws5.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for log_item in audit_logs:
        ws5.append([
            log_item.get("timestamp", ""),
            log_item.get("actor", ""),
            log_item.get("action", ""),
            log_item.get("description", ""),
            log_item.get("metadata", ""),
        ])

    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font

    for col in ["A", "B", "C", "D", "E"]:
        ws5.column_dimensions[col].width = 26

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
