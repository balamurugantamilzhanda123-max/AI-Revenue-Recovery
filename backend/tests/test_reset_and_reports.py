import io
import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.database import get_db
from app.main import app
from app.models import AuditLog, Customer, Order, PaymentStatus, RecoveryCase, Transaction
from app.services.reset_service import reset_operational_data
from app.services.report_service import generate_report_excel, generate_report_pdf, get_report_data


def test_reset_and_reports_flow(db_session: Session):
    client = TestClient(app)

    # 1. Create dummy customer, transaction, recovery case, and order
    cust = Customer(id="cust_test_reset_1", name="Reset Test Customer", email="reset.test@example.com")
    db_session.add(cust)
    db_session.flush()

    tx = Transaction(
        transaction_id="TXN-RESET-TEST-001",
        customer_id=cust.id,
        order_id="ORD-RESET-TEST-001",
        amount=15000.0,
        currency="INR",
        payment_method="UPI",
        status=PaymentStatus.FAILED,
        failure_reason="Network Connection Interrupted (TCP RST)",
    )
    db_session.add(tx)
    db_session.flush()

    order = Order(
        customer_id=cust.id,
        transaction_id=tx.id,
        product_id="prod_test_01",
        product_name="Test Router",
        category="Networking",
        unit_price=15000.0,
        subtotal=15000.0,
        total_amount=15000.0,
        status="PAYMENT_FAILED",
    )
    db_session.add(order)
    db_session.commit()

    # 2. Test dynamic report calculation
    report_data = get_report_data(db_session)
    assert report_data["summary"]["total_orders"] >= 1
    assert report_data["summary"]["revenue_at_risk"] >= 15000.0
    assert report_data["failure_analysis"]["network_errors"] >= 1
    assert len(report_data["transactions"]) >= 1

    # 3. Test PDF generation
    pdf_bytes = generate_report_pdf(report_data)
    assert len(pdf_bytes) > 100
    assert pdf_bytes.startswith(b"%PDF")

    # 4. Test Excel generation (4 sheets)
    excel_bytes = generate_report_excel(report_data)
    assert len(excel_bytes) > 100
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Summary" in wb.sheetnames
    assert "Transactions" in wb.sheetnames
    assert "Recovery Analysis" in wb.sheetnames
    assert "Risk Analysis" in wb.sheetnames

    # 5. Test API endpoints
    rep_res = client.get("/api/reports")
    assert rep_res.status_code == 200
    assert "summary" in rep_res.json()

    pdf_res = client.get("/api/reports/pdf")
    assert pdf_res.status_code == 200
    assert pdf_res.headers["content-type"] == "application/pdf"
    assert "attachment; filename=" in pdf_res.headers["content-disposition"]

    excel_res = client.get("/api/reports/excel")
    assert excel_res.status_code == 200
    assert "spreadsheetml" in excel_res.headers["content-type"]
    assert "attachment; filename=" in excel_res.headers["content-disposition"]

    # 6. Test reset endpoint
    reset_res = client.post("/api/admin/reset-dashboard")
    assert reset_res.status_code == 200
    assert reset_res.json()["success"] is True

    # 7. Verify all operational records are 0
    post_report = get_report_data(db_session)
    assert post_report["summary"]["total_orders"] == 0
    assert post_report["summary"]["revenue_at_risk"] == 0.0
    assert post_report["summary"]["total_recovered"] == 0.0
    assert post_report["summary"]["recovery_rate"] == 0.0
    assert len(post_report["transactions"]) == 0

    # 8. Verify Customer is preserved
    preserved_cust = db_session.get(Customer, "cust_test_reset_1")
    assert preserved_cust is not None

    # 9. Verify DASHBOARD_RESET audit log exists
    reset_log = db_session.query(AuditLog).filter(AuditLog.event_type == "DASHBOARD_RESET").first()
    assert reset_log is not None
    assert reset_log.event_message == "All operational transaction and revenue recovery data reset."
    assert "transactions_deleted" in reset_log.metadata_json
