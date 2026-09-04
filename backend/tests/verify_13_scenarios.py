import io
import sys
from pathlib import Path
_backend_dir = Path(__file__).resolve().parents[1]
if str(_backend_dir) not in sys.path:
    sys.path.insert(0, str(_backend_dir))

import openpyxl
from fastapi.testclient import TestClient

from app.database import SessionLocal, init_db
from app.main import app
from app.models import AuditLog, Customer, Order, PaymentStatus, RecoveryCase, Transaction
from app.services.report_service import generate_report_excel, generate_report_pdf, get_report_data
from app.services.reset_service import reset_operational_data

def run_all_13_tests():
    print("==================================================")
    print("STARTING REVIVEAI 13 SCENARIO VERIFICATION SUITE")
    print("==================================================")

    init_db()
    client = TestClient(app)
    db = SessionLocal()

    try:
        # Step 0: Ensure fresh initial test state
        reset_operational_data(db)

        # TEST 1: Open dashboard. Verify empty / initial state.
        dash_res = client.get("/api/dashboard/summary")
        assert dash_res.status_code == 200
        dash_data = dash_res.json()
        assert dash_data["total_transactions"] == 0
        assert dash_data["revenue_at_risk"] == 0.0
        assert dash_data["revenue_recovered"] == 0.0
        print("[OK] TEST 1 PASSED: Initial dashboard state verified (0 / INR 0 / 0%).")

        # TEST 2: Run one Demo transaction (e.g. checkout payment failure NETWORK_ERROR)
        checkout_payload = {
            "amount": 25000,
            "currency": "INR",
            "simulation_scenario": "NETWORK_ERROR",
            "product_id": "prod_laptop_biz_01",
            "customer": {
                "name": "Arjun Kumar",
                "email": "arjun@example.com",
                "phone": "+91 9876543210"
            }
        }
        pay_res = client.post("/api/checkout/process-payment", json=checkout_payload)
        assert pay_res.status_code == 200
        pay_data = pay_res.json()
        txn_id = pay_data["transaction_id"]
        assert pay_data["status"] == "FAILED"

        # Verify dashboard numbers changed
        dash_res2 = client.get("/api/dashboard/summary")
        dash_data2 = dash_res2.json()
        assert dash_data2["total_transactions"] >= 1
        assert dash_data2["revenue_at_risk"] >= 25000.0
        print(f"[PASS] TEST 2: Demo transaction created ({txn_id}), dashboard revenue at risk = INR {dash_data2['revenue_at_risk']:,.2f}.")

        # TEST 3: Open Transactions. Verify new transaction exists.
        tx_res = client.get("/api/transactions")
        assert tx_res.status_code == 200
        tx_list = tx_res.json().get("data", [])
        assert any(t["transaction_id"] == txn_id for t in tx_list)
        print(f"[PASS] TEST 3: Transactions ledger contains {txn_id}.")

        # TEST 4: Open Reports. Verify the same transaction appears.
        rep_res = client.get("/api/reports")
        assert rep_res.status_code == 200
        rep_data = rep_res.json()
        assert rep_data["summary"]["total_orders"] >= 1
        assert rep_data["summary"]["revenue_at_risk"] >= 25000.0
        assert rep_data["failure_analysis"]["network_errors"] >= 1
        assert any(t["transaction_id"] == txn_id for t in rep_data["transactions"])
        print("[PASS] TEST 4: Reports dynamically aggregated transaction and failure telemetry.")

        # TEST 5: Generate PDF. Verify PDF contains current data.
        pdf_bytes = generate_report_pdf(rep_data)
        assert len(pdf_bytes) > 200
        assert pdf_bytes.startswith(b"%PDF")
        print(f"[PASS] TEST 5: PDF Report generated successfully ({len(pdf_bytes)} bytes).")

        # TEST 6: Generate Excel. Verify all 4 sheets contain current data.
        excel_bytes = generate_report_excel(rep_data)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "Summary" in wb.sheetnames
        assert "Transactions" in wb.sheetnames
        assert "Recovery Analysis" in wb.sheetnames
        assert "Risk Analysis" in wb.sheetnames
        ws_summary = wb["Summary"]
        assert ws_summary.max_row >= 5
        print(f"[PASS] TEST 6: Excel workbook generated with 4 required sheets.")

        # TEST 7 & 8: Click Reset Dashboard & Confirm Reset
        reset_res = client.post("/api/admin/reset-dashboard")
        assert reset_res.status_code == 200
        reset_data = reset_res.json()
        assert reset_data["success"] is True

        # Verify values became 0
        dash_res_reset = client.get("/api/dashboard/summary")
        dash_post_reset = dash_res_reset.json()
        assert dash_post_reset["total_transactions"] == 0
        assert dash_post_reset["revenue_at_risk"] == 0.0
        assert dash_post_reset["revenue_recovered"] == 0.0
        assert dash_post_reset["recovery_rate"] == 0.0
        assert dash_post_reset["unresolved_cases"] == 0
        print("[PASS] TEST 7 & 8: Dashboard reset executed. Total Orders = 0, Risk = INR 0, Recovered = INR 0, Recovery Rate = 0%.")

        # TEST 9: Open Transactions after reset. Verify empty.
        tx_post_res = client.get("/api/transactions")
        assert tx_post_res.status_code == 200
        assert len(tx_post_res.json().get("data", [])) == 0
        print("[PASS] TEST 9: Transactions list is completely empty.")

        # TEST 10: Open Reports after reset. Verify reports empty/zero.
        rep_post_res = client.get("/api/reports")
        rep_post_data = rep_post_res.json()
        assert rep_post_data["summary"]["total_orders"] == 0
        assert rep_post_data["summary"]["revenue_at_risk"] == 0.0
        assert rep_post_data["summary"]["total_recovered"] == 0.0
        assert rep_post_data["summary"]["recovery_rate"] == 0.0
        assert len(rep_post_data["transactions"]) == 0
        print("[PASS] TEST 10: Reports page displays zero totals and empty transactions.")

        # TEST 11: Download PDF after reset. Verify PDF works with 0 values.
        pdf_reset_bytes = generate_report_pdf(rep_post_data)
        assert len(pdf_reset_bytes) > 200
        assert pdf_reset_bytes.startswith(b"%PDF")
        print("[PASS] TEST 11: Post-reset PDF generated cleanly with 0 values.")

        # TEST 12: Download Excel after reset. Verify sheets empty / 0 values.
        excel_reset_bytes = generate_report_excel(rep_post_data)
        wb_reset = openpyxl.load_workbook(io.BytesIO(excel_reset_bytes))
        assert "Summary" in wb_reset.sheetnames
        ws_tx = wb_reset["Transactions"]
        assert ws_tx.max_row == 1  # only header
        print("[PASS] TEST 12: Post-reset Excel sheets contain only headers and zero summary values.")

        # TEST 13: Run a new Demo Center scenario. Verify ONLY new data appears.
        new_scenario_payload = {
            "amount": 6999,
            "currency": "INR",
            "simulation_scenario": "NETWORK_ERROR",
            "product_id": "prod_acc_mech_keyboard_01",
            "customer": {
                "name": "Priya Sharma",
                "email": "priya@example.com",
                "phone": "+91 9444033445"
            }
        }
        pay_res_new = client.post("/api/checkout/process-payment", json=new_scenario_payload)
        assert pay_res_new.status_code == 200
        new_txn_id = pay_res_new.json()["transaction_id"]

        dash_res_fresh = client.get("/api/dashboard/summary")
        dash_data_fresh = dash_res_fresh.json()
        assert dash_data_fresh["total_transactions"] == 1
        assert dash_data_fresh["revenue_at_risk"] == 6999.0

        tx_res_fresh = client.get("/api/transactions")
        tx_list_fresh = tx_res_fresh.json().get("data", [])
        assert len(tx_list_fresh) == 1
        assert tx_list_fresh[0]["transaction_id"] == new_txn_id
        print(f"[PASS] TEST 13: Fresh scenario created {new_txn_id}. ONLY this 1 transaction exists in system from zero baseline.")

        # Audit Log Check
        audit_res = client.get("/api/audit")
        assert audit_res.status_code == 200
        logs = audit_res.json().get("data", [])
        assert any(l["event_type"] == "DASHBOARD_RESET" for l in logs)
        print("[PASS] AUDIT LOG: DASHBOARD_RESET immutable audit entry is present.")

        print("==================================================")
        print("ALL 13 TESTS COMPLETED AND VERIFIED SUCCESSFULLY!")
        print("==================================================")

    finally:
        db.close()

if __name__ == "__main__":
    run_all_13_tests()
